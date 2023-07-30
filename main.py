import requests
# import re

from bs4 import BeautifulSoup
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook


def get_wb():
    wb = load_workbook("result/DataExport.xlsx")
    return wb


def get_sheet(wb:load_workbook):
    sheet = wb['Chart data']
    return sheet


def check_link(link, mode, sheet, i, retries):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
        }
        r = requests.get(url=link, headers=headers)
        
        s = link.split("/")
        # l = len(s)
        if mode == 1:
            nameFile = s[2]
        if mode == 2:
            nameFile = s[3]
        
        with open(f"data/{nameFile}.html", "w", encoding="utf-8") as file:
            file.write(r.text)
        
        with open(f"data/{nameFile}.html", encoding="utf-8") as file:
            src = file.read()   
             
        soup = BeautifulSoup(src, "lxml")
        # comments = soup.find_all(class_=re.compile(r'\bgc-comment\b'))
        comments = soup.find_all(lambda tag: tag.has_attr('class') and 'gc-comment' in tag['class'])
        if len(comments) > 0:
            return True
        else:
            return False
    except requests.exceptions.SSLError as e:
        with open("log.txt", "a") as file:
            file.write(f"An SSL error occurred: {e}")
        print(f"An SSL error occurred: {e}")
        if retries > 0:
            with open("log.txt", "a") as file:
                file.write(f"Retrying ({retries} attempts left)...")
            print(f"Retrying ({retries} attempts left)...")
            return check_link(link, mode, sheet, i, retries=retries - 1)
        else:
            with open("log.txt", "a") as file:
                file.write("Max retries exceeded. Unable to fetch the link.")
            print("Max retries exceeded. Unable to fetch the link.")
            if mode == 1:
                sheet[f'H{i}'] = "bad site!"
            if mode == 2:
                sheet[f'I{i}'] = "bad site!"
            return False
    except Exception as e:
        with open("log.txt", "a") as file:
            file.write(f"An unexpected error occurred: {e}")
        print(f"An unexpected error occurred: {e}")
        if mode == 1:
            sheet[f'H{i}'] = "bad site!"
        if mode == 2:
            sheet[f'I{i}'] = "bad site!"
        return False


def get_data(sheet, wb):
    with open(f"result/log.txt", "w") as file:
        file.write("")

    i = 2
    while True:
        with open("log.txt", "a") as file:
            file.write(f"Row {i}\n")

        print(f"{i}")
        id = sheet[f'A{i}'].value 
        if id == None:
            break
        
        verificationLink = sheet[f'D{i}'].value 
        pageLink = sheet[f'E{i}'].value 

        if check_link(verificationLink, 1, sheet, i, 3):
            sheet[f'F{i}'] = "ok"
        else:
            sheet[f'F{i}'] = "bad"
        
        if check_link(pageLink, 2, sheet, i, 3):
            sheet[f'G{i}'] = "ok"
        else:
            sheet[f'G{i}'] = "bad"
        
        # if i == 100:
        #     wb.save("result/result.xlsx")
        #     break
        i += 1
    wb.save("result/result.xlsx")


def check_get_cource():
    wb = get_wb()
    sheet = get_sheet(wb)
    get_data(sheet, wb)

    
def main():
    print("start")
    check_get_cource()
    print("finish")


if __name__ == '__main__':
    main()
    
